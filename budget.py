import openpyxl as xlx
from openpyxl.chart import BarChart, Reference


def budget_workbook(file_name):
    wb = xlx.load_workbook(file_name)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        income_cell = sheet.cell(row, 2)

        amount_to_put_in_savings = income_cell.value * .15
        savings_cell = amount_to_put_in_savings
        savings_cell.value = sheet.cell(row, 3)

        amount_to_put_in_401k = income_cell.value * .15
        retirement_cell = amount_to_put_in_401k
        retirement_cell.value = sheet.cell(row, 4)

        amount_to_give_to_charity = income_cell.value * .1
        charity_cell = amount_to_give_to_charity
        charity_cell.value = sheet.cell(row, 5)

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=3, max_col=5)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "h2")

    wb.save(file_name)
